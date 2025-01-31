######################################
# Code for building the spreadsheet
######################################
from metrics_utility.automation_controller_billing.report.base import Base
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

class ReportCCSP(Base):
    BLACK_COLOR_HEX = "00000000"
    WHITE_COLOR_HEX = "00FFFFFF"
    BLUE_COLOR_HEX = "000000FF"
    GREEN_COLOR_HEX = "0000FF00"
    FONT = "Arial"
    PRICE_FORMAT = '$#,##0.00'

    def __init__(self, dataframe, report_period, extra_params):
        # Create the workbook and worksheet
        self.wb = Workbook()

        self.dataframe = dataframe
        self.report_period = report_period
        self.extra_params = extra_params
        self.price_per_node = extra_params['price_per_node']

        self.config = {
            'sku': extra_params['report_sku'],
            'h1_heading': {
                'value': extra_params['report_h1_heading'],
            },
            'h2_heading': {
                'value': 'Monthly payment',
            },
            'header': [{
                'label': 'CCSP Company Name',
                'value': extra_params['report_company_name'],
            },{
                'label': 'CCSP Email',
                'value': extra_params['report_email'],
            },{
                'label': 'CCSP RHN Login',
                'value': extra_params['report_rhn_login'],
            },{
                'label': 'Report Period (YYYY-MM)',
                'value': '<autogenerated>',
            },{
                'label': 'Company Business leader ',
                'value': extra_params['report_company_business_leader'],
            },{
                'label': 'Company Procurement leader ',
                'value': extra_params['report_company_procurement_leader'],
            },{
                'label': 'Periodicity',
                'value': 'Report is submitted monthly',
            }],
        }

        default_sku_description = [
            ['SKU',
            'SKU Description',
            '',
            'Term',
            'Unit of Measure',
            'Currency',
            'MSRP'],
            [f"{extra_params['report_sku']}",
            f"{extra_params['report_sku_description']}",
            '',
            'MONTH',
            'MANAGED NODE',
            'USD',
            f"{extra_params['price_per_node']}"]
        ]

        default_column_widths = {
            1: 63,
            2: 63,
            3: 15,
            4: 15,
            5: 15,
        }

        default_data_column_widths = {
            1: 40,
            2: 20,
            3: 20,
            4: 20,
            5: 20,
            6: 20,
            7: 20
        }

        self.config['sku_description'] = default_sku_description
        self.config['column_widths'] = default_column_widths
        self.config['data_column_widths'] = default_data_column_widths

    def build_spreadsheet(self):
        job_host_sumary_dataframe = self.dataframe[0]
        events_dataframe = self.dataframe[1]
        events_dataframe = self._fix_event_host_names(job_host_sumary_dataframe, events_dataframe)

        # Create the workbook and worksheets
        self.wb.remove(self.wb.active) # delete the default sheet
        self.wb.create_sheet(title="Usage Reporting")

        # First sheet with billing
        ws = self.wb.worksheets[0]

        self._init_dimensions(ws)
        current_row = self._build_heading_h1(1, ws)
        current_row = self._build_header(current_row, ws)
        current_row = self._build_heading_h2(current_row, ws)
        current_row = self._build_sku_description(current_row, ws)
        current_row = self._build_data_section(current_row, ws, job_host_sumary_dataframe)

        # Add optional sheets
        sheet_index = 1
        if "managed_nodes" in self.optional_report_sheets():
            # Sheet with list of managed nodes
            self.wb.create_sheet(title="Managed nodes")
            ws = self.wb.worksheets[sheet_index]
            current_row = self._build_data_section_usage_by_node(1, ws, job_host_sumary_dataframe)
            sheet_index += 1

        if events_dataframe is not None:
            if "usage_by_collections" in self.optional_report_sheets():
                # Sheet with usage by collections
                self.wb.create_sheet(title="Usage by collections")
                ws = self.wb.worksheets[sheet_index]
                current_row = self._build_data_section_usage_by_collections(1, ws, events_dataframe)
                sheet_index += 1

            if "usage_by_roles" in self.optional_report_sheets():
                # Sheet with usage by roles
                self.wb.create_sheet(title="Usage by roles")
                ws = self.wb.worksheets[sheet_index]
                current_row = self._build_data_section_usage_by_roles(1, ws, events_dataframe)
                sheet_index += 1

            if "usage_by_modules" in self.optional_report_sheets():
                # Sheet with usage by modules
                self.wb.create_sheet(title="Usage by modules")
                ws = self.wb.worksheets[sheet_index]
                current_row = self._build_data_section_usage_by_modules(1, ws, events_dataframe)
                sheet_index += 1

        return self.wb

    def _init_dimensions(self, ws):
        for key, value in self.config['column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

    def _build_heading_h1(self, current_row, ws):
        # Merge cells and insert the h1 heading
        ws.merge_cells(start_row=current_row, start_column=1,
                            end_row=current_row, end_column=2)
        h1_heading_cell = ws.cell(row=current_row, column=1)
        h1_heading_cell.value = self.config['h1_heading']['value']

        h1_heading_cell.fill = PatternFill("solid", fgColor=self.BLACK_COLOR_HEX)
        h1_heading_cell.font = Font(name=self.FONT,
                                    size=17,
                                    color=self.WHITE_COLOR_HEX)

        h1_heading_cell.alignment = Alignment(horizontal='center')

        current_row += 1
        return current_row

    def _build_heading_h2(self, current_row, ws):
        # Add the h2 heading payment heading
        black_background = PatternFill("solid", fgColor=self.BLACK_COLOR_HEX)
        green_background = PatternFill("solid", fgColor=self.GREEN_COLOR_HEX)

        # black background row
        ws.cell(row=current_row, column=1).fill = black_background
        ws.cell(row=current_row, column=2).fill = black_background
        current_row += 1

        # h2 heading with green background
        cell = ws.cell(row=current_row, column=1)
        cell.value = self.config['h2_heading']['value']
        cell.fill = green_background
        ws.cell(row=current_row, column=2).fill = green_background
        current_row += 1

        # black background row
        ws.cell(row=current_row, column=1).fill = black_background
        ws.cell(row=current_row, column=2).fill = black_background
        current_row += 1

        # make extra blank 2 rows
        current_row += 2

        return current_row

    def _build_header(self, current_row, ws):
        # Insert the header
        for header_row in self.config['header']:
            header_label_font = Font(name=self.FONT,
                                     size=14,
                                     color=self.WHITE_COLOR_HEX)
            header_value_font = Font(name=self.FONT,
                                     size=14,
                                     color=self.BLACK_COLOR_HEX)
            header_label_fill = PatternFill("solid", fgColor=self.BLUE_COLOR_HEX)

            cell = ws.cell(row=current_row, column=1)
            cell.value = header_row['label']
            cell.font = header_label_font
            cell.fill = header_label_fill

            cell = ws.cell(row=current_row, column=2)
            if header_row['label'] == "Report Period (YYYY-MM)":
                # Insert dynamic report period into the specific header column
                cell.value = self.report_period
            else:
                cell.value = header_row['value']
            cell.font = header_value_font
            current_row += 1

        return current_row

    def _build_sku_description(self, current_row, ws):
        # Insert the header
        row_counter = 0
        for header_row in self.config['sku_description']:
            header_font = Font(name=self.FONT,
                               size=11,
                               color=self.BLACK_COLOR_HEX,
                               bold=True)
            header_border = Border(left=Side(border_style='thin',
                                             color=self.BLACK_COLOR_HEX),
                                   right=Side(border_style='thin',
                                              color=self.BLACK_COLOR_HEX),
                                   top=Side(border_style='thin',
                                            color=self.BLACK_COLOR_HEX),
                                   bottom=Side(border_style='thin',
                                               color=self.BLACK_COLOR_HEX))
            value_font = Font(name=self.FONT,
                              size=11,
                              color=self.BLACK_COLOR_HEX)
            col_counter = 0
            for col_value in header_row:
                col_counter += 1

                cell = ws.cell(row=current_row + row_counter, column=col_counter)
                cell.value = col_value

                if row_counter == 0:
                    # header
                    cell.font = header_font
                    cell.border = header_border
                else:
                    # row
                    cell.font = value_font

            row_counter +=1
        current_row = current_row + row_counter
        # make extra 1 row space
        current_row += 1

        return current_row

    def _build_data_section(self, current_row, ws, dataframe):
        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        dotted_border = Border(left=Side(border_style='dotted',
                                         color=self.BLACK_COLOR_HEX),
                               right=Side(border_style='dotted',
                                          color=self.BLACK_COLOR_HEX),
                               top=Side(border_style='dotted',
                                        color=self.BLACK_COLOR_HEX),
                               bottom=Side(border_style='dotted',
                                           color=self.BLACK_COLOR_HEX))

        ccsp_report = dataframe.reset_index().groupby(
            'organization_name', dropna=False).agg(
                quantity_consumed=('host_name', 'nunique'))
        ccsp_report['mark_x'] = ''
        ccsp_report['unit_price'] = round(self.price_per_node, 2)
        ccsp_report['extended_unit_price'] = round((ccsp_report['quantity_consumed'] * ccsp_report['unit_price']), 2)

        # order the columns right
        ccsp_report = ccsp_report.reset_index()
        ccsp_report = ccsp_report.reindex(columns=['organization_name',
                                                   'mark_x',
                                                   'quantity_consumed',
                                                   'unit_price',
                                                   'extended_unit_price'])

        # Rename the columns based on the template
        ccsp_report_dataframe = ccsp_report.rename(
            columns={"organization_name": "Organization name (i.e. company name)",
                     "mark_x": "Please Mark With An 'X' If The Usage Is Internal. \nOtherwise Leave Blank",
                     "quantity_consumed": "Red Hat SKU\n Quantity Consumed",
                     "unit_price": "Subscription Fee\n (SKU Unit Price)",
                     "extended_unit_price": "Extended\n Subscription Fees\n (SKU Extended Unit Price)"
                    })

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                cell.border = dotted_border
                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                else:
                    # set value style
                    cell.font = value_font
                    if c_idx >= 4:
                        # Format all price cols
                        cell.number_format = self.PRICE_FORMAT
                    if c_idx == 5:
                        # Override the value of the extended price (number of nodes X unitp rice)
                        # Multiply columns 3x4 instead of inserting the price per org
                        cell_m_1 = ws.cell(row=r_idx, column=3).column_letter + str(r_idx)
                        cell_m_2 = ws.cell(row=r_idx, column=4).column_letter + str(r_idx)
                        cell.value = '={0}*{1}'.format(cell_m_1, cell_m_2)

            row_counter += 1

        if row_counter >= 2:
            # If there is at least 1 data column, insert the sum:
            first_row = current_row + 1 # ignore the header
            last_row = current_row + row_counter - 1
            sum_row = current_row + row_counter

            # Sum description
            cell = ws.cell(row=sum_row, column=1)
            cell.value = "Sum of monthly payment"

            # Sum value
            cell = ws.cell(row=sum_row, column=5)
            cell_sum_start = cell.column_letter + str(first_row)
            cell_sum_end = cell.column_letter + str(last_row)
            cell.value = '=SUM({0}:{1})'.format(cell_sum_start, cell_sum_end)

        return current_row + row_counter + 1 # include the sum row after